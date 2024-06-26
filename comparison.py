import base64
import io
import os
import tempfile
from openpyxl import Workbook, load_workbook
import pandas as pd
import streamlit as st
from pathlib import Path
from io import BytesIO
from PIL import Image
import cv2
from utils_app import create_fig, load_model
import config
#from app import load_model
import numpy as np
import re
from yolov7.yolov7_wrapper import YOLOv7Wrapper
import datetime
import json
import time

  
@st.cache_data()
def detect_with_v7(uploaded_file,model_name, confidence_threshold=0.5):
    # Check if an uploaded_file is provided
    if not uploaded_file:
        return None, None

    # Load the image and convert it to a numpy array
    uploaded_img_data = uploaded_file.getvalue()
    im0 = Image.open(BytesIO(uploaded_img_data)).convert('RGB')
    im0_np = np.array(im0)

    # Initialize a dictionary to store results from all models
    model_results = {}

    
    # st.write(f" model name before the wrapper : {model_name}") debugging msg
    yolov7_model = YOLOv7Wrapper(model_name)
    # Perform detection and get back the image and captions (detections)
    detected_image, captions = yolov7_model.detect_and_draw_boxes_from_np(im0_np, confidence_threshold=confidence_threshold)

    confidence_threshold = 0.5  

    boxes = {'count': {}, 'details': ''}
    for caption in captions:
        parts = caption.split()
        class_name = parts[0].split('=')[1]
        
        # Initialize default values for box and confidence
        box = "Unavailable"
        confidence = 0.0

        # Check if 'parts' has enough elements to extract coordinates and confidence
        if len(parts) >= 5:
            # Attempt to extract the coordinates
            try:
                # The coordinates are expected to be in parts[1], parts[2], parts[3], and parts[4]
                x_min = parts[1].split('=')[1].strip('()').strip(',')
                y_min = parts[2].strip(',')
                x_max = parts[3].strip(',')
                y_max = parts[4].strip(')')
                box = f"{x_min}, {y_min}, {x_max}, {y_max}"
            except Exception as e:
                print(f"Error parsing coordinates: {e}")
                box = "Unavailable"

            # Attempt to extract the confidence
            try:
                confidence_part = next(part for part in parts if 'confidence' in part)
                confidence = float(confidence_part.split('=')[1].rstrip('%'))
            except Exception as e:
                print(f"Error parsing confidence: {e}")
                confidence = 0.0

        # Update dictionary only if confidence is above threshold
        if confidence >= confidence_threshold * 100:  # Assuming confidence_threshold is in [0, 1]
            if class_name not in boxes['count']:
                boxes['count'][class_name] = {'count': 0, 'entries': []}
            entry = {'coordinates': box, 'confidence': confidence}
            boxes['count'][class_name]['entries'].append(entry)
            boxes['count'][class_name]['count'] += 1





        
        

    
    model_results[model_name] = boxes

    
    results = {}
    for model_name, boxes in model_results.items():
        count_dict = {}
        detection_results = ""  # This ensures we start with a clean slate for the detection results

        # Iterate over the 'count' dictionary inside the 'boxes' dictionary
        for class_name, details in boxes['count'].items():
            for entry in details['entries']:
                
                
                # Append detection info to the results string in the specified format
                detection_results += f"<b style='color: blue;'>Object type:</b> {class_name}<br>"
                detection_results += f"<b style='color: blue;'>Coordinates:</b> {entry['coordinates']}<br>"
                detection_results += f"<b style='color: blue;'>Confidence:</b> {entry['confidence']}%<br>---<br>"
                
            # Populate the count dictionary for each class detected
            if class_name not in count_dict:
                count_dict[class_name] = {'count': details['count']}

        # Assign the structured results to the results dictionary under the current model name
        results[model_name] = {
            "count": count_dict,
            "details": detection_results
        }

    
    if isinstance(detected_image, Image.Image):
        detected_pil_image = detected_image
    else:
        # This will handle the case where detected_image might be a numpy array (or any other type)
        
        try:
            detected_pil_image = Image.fromarray(detected_image.astype(np.uint8))
        except Exception as e:
            print(f"Error converting detected_image to PIL image: {e}")
            detected_pil_image = None  # Or you can provide a default image here

    

    # Explicitly delete large objects and free memory
    # del model
    #gc.collect()
    
    return detected_pil_image, results

  
def detect_with_v8(uploaded_file, model, conf=0.5):
    """
    Execute inference for uploaded image with YOLOv8 model.

    Parameters:
    - uploaded_file: The uploaded image file.
    - model: An instance of the `YOLOv8` class containing the YOLOv8 model.
    - conf: Confidence threshold for YOLOv8 model.

    Returns:
    - detected_image: Processed image with detections (PIL Image).
    - results: Dictionary containing detection results.
    """

    if uploaded_file is None:
        return None, None

    try:
        uploaded_image = Image.open(uploaded_file)
    except Exception as e:
        print(f"Error opening uploaded image: {e}")
        return None, None
    
    detected_image_result = model.predict(uploaded_image, conf=conf)
    boxes = detected_image_result[0].boxes

    # Get the plotted image with detections from the Results object
    try:
        detected_img_arr = detected_image_result[0].plot()[:, :, ::-1]  # Assuming this returns a numpy array
        # Convert the numpy array to a PIL Image object
        detected_image = Image.fromarray(cv2.cvtColor(detected_img_arr, cv2.COLOR_BGR2RGB))
    except Exception as e:
        print(f"Error processing detected image: {e}")
        detected_image = None

    detection_results = ""
    count_dict = {}
    for box in boxes:
        class_id = model.names[box.cls[0].item()]
        cords = box.xyxy[0].tolist()
        cords = [round(x) for x in cords]
        conf = round(box.conf[0].item(), 2)

        detection_results += f"<b style='color: blue;'>Object type:</b> {class_id}<br><b style='color: blue;'>Coordinates:</b> {cords}<br><b style='color: blue;'>Confidence:</b> {conf}<br>---<br>"
        if class_id in count_dict:
            count_dict[class_id] += 1
        else:
            count_dict[class_id] = 1

    results = {
        "count": count_dict,
        "details": detection_results
    }
      # Display the results for debugging.
    #model_architecture = str(model)
    
    return detected_image, results #, model_architecture
    
def run_detection(detection_model_name, file, confidence):        
            # Detect based on the model
            
            if detection_model_name in config.DETECTION_MODEL_DICT_V7:
                
                    with st.spinner(f'Detecting objects with {detection_model_name}...'):  # YOLOv7
                        model_path = config.DETECTION_MODEL_DICT_V7[detection_model_name]
                        
                        detected_image, results = detect_with_v7(file, detection_model_name, confidence_threshold=confidence)
                        
            else:  # YOLOv8
                model_path = config.DETECTION_MODEL_DICT_V8[detection_model_name] 
                model_instance = load_model(model_path) 
                if not model_instance:
                    st.write(f"Error: Failed to load YOLOv8 model {detection_model_name}")
                detected_image, results = detect_with_v8(file, model_instance, conf=confidence)
            
           
            
            return detected_image, results #, model_architecture#
                    

def compare_models_function():
    st.markdown("### Compare different models' performance")
    st.divider()
    

    
    st.sidebar.divider()
    # New UI for YOLOv7 models
    st.sidebar.markdown("### :camera: YOLOv7 Models")
    selected_models_v7 = st.sidebar.multiselect(
        "Select YOLOv7 models for comparison:",
        config.DETECTION_MODEL_LIST_V7, 
        key="models_comparison_selectbox_v7"
    )

    # UI for other models (V8)
    st.sidebar.markdown("### :rocket: YOLOv8 Models")
    selected_models_v8 = st.sidebar.multiselect(
        "Select other models for comparison:",
        config.DETECTION_MODEL_LIST_V8,
        key="models_comparison_selectbox_v8"
    )
    
    selected_models = selected_models_v7 + selected_models_v8
    # Check if any models are selected
    if not selected_models:
        st.info("Please select models for comparison.")
        
        return
    conf = float(st.sidebar.slider(
        "Select Model Confidence", 30, 100, 50)) / 100

    if len(selected_models) > 4:
        st.warning("You can only select up to 4 models.")
        return

    uploaded_file = st.sidebar.file_uploader("Choose an image for comparison...", type=["jpg", "png"])

    animation_placeholder = st.empty()

    if uploaded_file is not None and uploaded_file.type.startswith('image/'):
        st.sidebar.divider()
        st.sidebar.image(uploaded_file)
        # Display the animation
        with animation_placeholder.container():
            display_animation()
    
            
    
    #------- Layout for displaying images in 2x2 grid
    
    
    if len(selected_models) > 2:
        # Create a 2x2 grid for four models
        row1_col1, row1_col2 = st.columns([0.45, 0.45])
        row2_col1, row2_col2 = st.columns([0.45, 0.45])
        column_layouts = [(row1_col1, row1_col2), (row2_col1, row2_col2)]
    else:
        #-------- Create a single row for fewer models
        column_layouts = [st.columns(2) for _ in range(len(selected_models))]

    if uploaded_file:
        detected_images = {}
        all_results = {}
        aggregated_results = {} 
        inference_times = {}

        # Automatically run detection for each selected model
        for idx, selected_model_name in enumerate(selected_models):
            if idx > 3:
                break

            # Run detection
            start_time = time.time()
            detected_image, results = run_detection(selected_model_name, uploaded_file, conf)
            end_time = time.time()
            
            # Calculate and store the inference time for the current model
            inference_time = end_time - start_time
            inference_times[selected_model_name] = inference_time

            # Store detection results
            detected_images[selected_model_name] = detected_image
            all_results[selected_model_name] = results
            


        # Aggregate results
        for selected_model_name, results in all_results.items():
            
            if selected_model_name in config.DETECTION_MODEL_LIST_V7:
                update_v7_results(aggregated_results, results, selected_model_name)
            else:
                update_v8_results(aggregated_results, results, selected_model_name)


        animation_placeholder.empty()


        # Display detected images side by side
        for idx, (model_name, detected_image) in enumerate(detected_images.items()):
            row_idx = idx // 2  # Row index: 0 or 1
            col_idx = idx % 2   # Column index: 0 or 1
            with column_layouts[row_idx][col_idx]:
                if detected_image:
                    fig = create_fig(detected_image, detected=True)
                    st.plotly_chart(fig, use_container_width=True)
                    st.caption(model_name)

                    # Display inference time
                    if model_name in inference_times:
                        st.write(f"Inference time: {inference_times[model_name]:.2f} seconds")

        st.write("### Comparison Results")
        st.table(aggregated_results)
        
        display_model_results(all_results)
        st.divider()
        #print(all_results)
        if st.button('Save Results'):
                data_to_save = []
                detailed_results = {}

                # Process results from each model
                for model_name, model_data in all_results.items():
                    aggregated_results, detailed_results = parse_and_store_details(
                        model_name, model_data, aggregated_results, detailed_results
                    )
                    # print(detailed_results)
                    # print(aggregated_results)
                    #print(model_data)
                    # Prepare data for each detection in detailed_results
                    for class_id, detections in detailed_results.items():
                        for detection in detections:
                            if detection['model'] == model_name:  # Ensure the detection belongs to the current model
                                detection_entry = {
                                    "model": model_name,
                                    "inference_details": {
                                        "class_id": class_id,
                                        "count": aggregated_results.get(class_id, {}).get(model_name, 1),  # Get count from aggregated_results
                                        "coordinates": detection['coordinates'],
                                        "confidence": detection['confidence']
                                    }
                                }
                                data_to_save.append(detection_entry)
                                print(detection_entry)

               
                #print(data_to_save)
                

                # Now, save data to Firebase
                # try:
                #     save_to_firebase(data_to_save, user_id)
                # except Exception as e:
                #     st.error(f" to Firebase: {e}")
                #save_data(data_to_save)
                save_results(data_to_save)

def save_results(data_to_save):
    
    
    # data = []
    # for entry in data_to_save:
    #     row = {
    #         "model": entry["model"],
    #         "class_id": entry["inference_details"]["class_id"],
    #         "count": entry["inference_details"]["count"],
    #         "coordinates": entry["inference_details"]["coordinates"],
    #         "confidence": entry["inference_details"]["confidence"]
    #     }
    #     data.append(row)

    # df = pd.DataFrame(data)

    # # Save the DataFrame to a temporary file
    # results_file = os.path.join(os.path.dirname(__file__), "results.xlsx")
    # df.to_excel(results_file, index=False)

    # st.success(f"Detected images and results saved to 'results.xlsx' in the project folder.")
    data = []
    for entry in data_to_save:
        current_timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row = {
            "timestamp": current_timestamp,
            "model": entry["model"],
            "class_id": entry["inference_details"]["class_id"],
            "count": entry["inference_details"]["count"],
            "coordinates": entry["inference_details"]["coordinates"],
            "confidence": entry["inference_details"]["confidence"]
        }
        data.append(row)
    df = pd.DataFrame(data)

    # Save the DataFrame to an Excel file in the project folder
    results_file = os.path.join(os.path.dirname(__file__), "results.xlsx")

    # Check if the file already exists
    if os.path.exists(results_file):
        # If the file exists, append the new data to the existing sheet
        book = load_workbook(results_file)
        sheet = book.active
        row_count = sheet.max_row + 1
        for _, row in df.iterrows():
            sheet.cell(row=row_count, column=1, value=row['timestamp'])
            sheet.cell(row=row_count, column=2, value=row['model'])
            sheet.cell(row=row_count, column=3, value=row['class_id'])
            sheet.cell(row=row_count, column=4, value=row['count'])
            sheet.cell(row=row_count, column=5, value=row['coordinates'])
            sheet.cell(row=row_count, column=6, value=row['confidence'])
            row_count += 1
        book.save(results_file)
    else:
        # If the file doesn't exist, create a new file
        book = Workbook()
        sheet = book.active
        sheet.title = "Results"
        sheet.cell(row=1, column=1, value="timestamp")
        sheet.cell(row=1, column=2, value="model")
        sheet.cell(row=1, column=3, value="class_id")
        sheet.cell(row=1, column=4, value="count")
        sheet.cell(row=1, column=5, value="coordinates")
        sheet.cell(row=1, column=6, value="confidence")
        row_count = 2
        for _, row in df.iterrows():
            sheet.cell(row=row_count, column=1, value=row['timestamp'])
            sheet.cell(row=row_count, column=2, value=row['model'])
            sheet.cell(row=row_count, column=3, value=row['class_id'])
            sheet.cell(row=row_count, column=4, value=row['count'])
            sheet.cell(row=row_count, column=5, value=row['coordinates'])
            sheet.cell(row=row_count, column=6, value=row['confidence'])
            row_count += 1
        book.save(results_file)

    st.success(f"Detected images and results saved to 'results.xlsx' in the project folder.")

def save_data(detected_images, all_results):
    # Convert the PIL Image objects to base64-encoded strings
    detected_images_data = []
    for model_name, detected_image in detected_images.items():
        if detected_image:
            buffered = io.BytesIO()
            detected_image.save(buffered, format="PNG")
            detected_images_data.append((model_name, base64.b64encode(buffered.getvalue()).decode("utf-8")))

    data = {
        "detected_images": detected_images_data,
        "all_results": all_results,
        "timestamp": datetime.datetime.now().isoformat()
    }

    # Save the data to a file in the project folder
    results_file = os.path.join(os.path.dirname(__file__), "results.txt")
    try:
        with open(results_file, "w") as f:
            json.dump(data, f, indent=4)
        st.success("Detected images and results saved to 'results.txt' in the project folder.")
    except Exception as e:
        st.error(f"Error saving data to 'results.txt': {e}")

def save_to_files(detected_images, all_results):
    # data = []
    # for model_name, detected_image in detected_images.items():
    #     if detected_image:
    #         buffered = io.BytesIO()
    #         detected_image.save(buffered, format="PNG")
    #         image_data = buffered.getvalue()
    #         row = {
    #             "model_name": model_name,
    #             "image_data": image_data
    #         }
    #         data.append(row)

    # for model_name, result in all_results.items():
    #     if isinstance(result, dict):
    #         for class_id, detections in result.get("count", {}).items():
    #             if isinstance(detections, dict):
    #                 for detection in detections.get("entries", []):
    #                     row = {
    #                         "model": model_name,
    #                         "class_id": class_id,
    #                         "count": detections.get("count", 1),
    #                         "coordinates": detection.get("coordinates"),
    #                         "confidence": detection.get("confidence")
    #                     }
    #                     data.append(row)
    #             elif isinstance(detections, int):
    #                 row = {
    #                     "model": model_name,
    #                     "class_id": class_id,
    #                     "count": detections,
    #                     "coordinates": "Unavailable",
    #                     "confidence": 0.0
    #                 }
    #                 data.append(row)
    #             else:
    #                 st.warning(f"Unexpected 'detections' format for model {model_name}, class_id {class_id}: {detections}")
    #     else:
    #         st.warning(f"Unexpected 'result' format for model {model_name}: {result}")

    # df = pd.DataFrame(data)
    # results_file = os.path.join(os.path.dirname(__file__), "results.xlsx")
    # df.to_excel(results_file, index=False)

    # st.success("Detected images and results saved to 'results.xlsx' in the project folder.")

    data = []
    for model_name, detected_image in detected_images.items():
        if detected_image:
            buffered = io.BytesIO()
            detected_image.save(buffered, format="PNG")
            image_data = buffered.getvalue()
            row = {
                "model_name": model_name,
                "image_data": image_data
            }
            data.append(row)

    for model_name, result in all_results.items():
        if model_name in config.DETECTION_MODEL_LIST_V7:
            # Process YOLOv7 results
            count_dict = result.get("count", {})
            for class_id, count_details in count_dict.items():
                for entry in count_details.get("entries", []):
                    row = {
                        "model": model_name,
                        "class_id": class_id,
                        "count": count_details["count"],
                        "coordinates": entry["cords"],
                        "confidence": entry["conf"]
                    }
                    data.append(row)
        elif model_name in config.DETECTION_MODEL_LIST_V8:
            # Process YOLOv8 results
            count_dict = result.get("count", {})
            details_dict = result.get("details", {})
            for class_id, count_value in count_dict.items():
                cords = details_dict.get(class_id, {}).get("cords", "Unavailable")
                conf = details_dict.get(class_id, {}).get("conf", 0.0)
                row = {
                    "model": model_name,
                    "class_id": class_id,
                    "count": count_value,
                    "coordinates": cords,
                    "confidence": conf
                }
                data.append(row)
        else:
            st.warning(f"Unexpected model name: {model_name}")

    df = pd.DataFrame(data)
    results_file = os.path.join(os.path.dirname(__file__), "results.xlsx")
    df.to_excel(results_file, index=False)

    st.success("Detected images and results saved to 'results.xlsx' in the project folder.")


def display_animation():
    lottie = """
            <script src="https://unpkg.com/@lottiefiles/lottie-player@latest/dist/lottie-player.js"></script>
            <lottie-player src="animation_sphere.json" background="transparent" speed="1" style="width: 800px; height: 800px;" loop autoplay></lottie-player>
            """
    st.markdown("""
        <style>
            iframe {
                position: fixed;
                top: 16rem;
                bottom: 0;
                left: 205;
                right: 0;
                margin: auto;
                z-index=-1;
            }
        </style>
        """, unsafe_allow_html=True
    )

    st.write("🔍 Analyzing your image... Hang tight, awesome detections are on their way! ")
    st.components.v1.html(lottie, width=810, height=810)        

            
def update_v7_results(aggregated_results, results, model_name):
    

    # Extract the 'count' dictionary from the results
    count_dict = results.get(model_name, {}).get('count', {})
    if count_dict:
        for class_id, count_details in count_dict.items():
            if class_id not in aggregated_results:
                aggregated_results[class_id] = {}
            aggregated_results[class_id][model_name] = count_details['count']
    # else:
    #     st.write(f"No 'count' found in results for model: {model_name}")
 


def update_v8_results(aggregated_results, results, model_name):
    count = results.get("count")
    # Assuming `details` contains coordinates for each class_id
    

    if count:
        for class_id, count_value in count.items():
            if class_id not in aggregated_results:
                aggregated_results[class_id] = {}

            aggregated_results[class_id][model_name] = count_value
                
def display_model_results(model_results):
    # Grid layout: 2x2 for up to four models
    columns = [st.columns(2), st.columns(2)]

    for idx, (model_name, results) in enumerate(model_results.items()):
        with columns[idx // 2][idx % 2]:
            if model_name in config.DETECTION_MODEL_LIST_V7:
                handle_v7_results(results, model_name)
            else:
                handle_v8_results(results, model_name)          

def handle_v7_results(results, model_name):
    with st.expander(f"Detailed results for {model_name}"):
        for model_name, result in results.items():  # Corrected iteration
            if result["details"].strip():
                scrollable_textbox = f"""
                    <div style="font-family: 'Source Code Pro','monospace'; font-size: 16px; overflow-y: scroll; border: 1px solid #000; padding: 10px; width: 500px; height: 400px;">
                        {result["details"]}
                    </div>
                """
                st.markdown(scrollable_textbox, unsafe_allow_html=True)
            else:
                st.markdown(f"No objects detected by {model_name}. Please try a different image or adjust the model's confidence threshold.")





def handle_v8_results(results, model_name):

    if not isinstance(results, dict):
        raise ValueError("Expected results to be a dictionary.")

    
    details = results.get("details", "").strip()

    
    with st.expander(f"Detailed results for {model_name}"):
        # Detailed results display
        if details:
            
            scrollable_textbox = f"""
            <div style="
                font-family: 'Source Code Pro','monospace';
                font-size: 16px;
                overflow-y: scroll;
                border: 1px solid #000;
                padding: 10px;
                width: 500px;
                height: 400px;
            ">
                {details}
            </div>
            """
            st.markdown(scrollable_textbox, unsafe_allow_html=True)
        else:
            
            st.markdown(f"No objects detected by {model_name}.Please try a different image or adjust the model's confidence threshold.")


    


def parse_and_store_details(model_name, results, aggregated_results, detailed_results):
    # Check if the model is YOLOv7
    if model_name in config.DETECTION_MODEL_LIST_V7:
        count_dict = results.get(model_name, {}).get('count', {})
        if count_dict:
            for class_id, count_details in count_dict.items():
                if class_id not in aggregated_results:
                    aggregated_results[class_id] = {}
                aggregated_results[class_id][model_name] = count_details['count']

        details_str = results.get(model_name, {}).get('details', {})
        # st.write("details_str:", details_str)

        

        detections = details_str.split('---')
        
        for detection in detections:
            parts = detection.split('<br>')
            class_id, coordinates, confidence = None, None, None
            for part in parts:
                if 'Object type:' in part:
                    # Extracting class_id while removing HTML tags
                    class_id_match = re.search(r"Object type:</b>\s*(\w+)", part)
                    if class_id_match:
                        class_id = class_id_match.group(1)
                    else:
                        print("Error parsing class ID:", part)
                        class_id = None
                elif 'Coordinates:' in part:
                    # Extracting coordinates while removing HTML tags
                    coordinates_match = re.search(r"Coordinates:</b>\s*(.+)", part)
                    if coordinates_match:
                        coordinates = coordinates_match.group(1)
                    else:
                        print("Error parsing coordinates:", part)
                        coordinates = None
                elif 'Confidence:' in part:
                    # Using regular expression to extract the numeric confidence value
                    match = re.search(r'Confidence:</b>\s*(\d+(\.\d+)?)', part)
                    if match:
                        confidence_str = match.group(1)
                        confidence = float(confidence_str) / 100
                    else:
                        print("Error parsing confidence value:", part)
                        confidence = None

            if class_id and coordinates and confidence is not None:
                if class_id not in detailed_results:
                    detailed_results[class_id] = []
                detailed_results[class_id].append({
                    'model': model_name,
                    'coordinates': coordinates,
                    'confidence': confidence
                })


    # Check if the model is YOLOv8
    elif model_name in config.DETECTION_MODEL_LIST_V8:
        count = results.get("count", {})
        if count:
            for class_id, count_value in count.items():
                if class_id not in aggregated_results:
                    aggregated_results[class_id] = {}
                aggregated_results[class_id][model_name] = count_value
     # Process detailed results for YOLOv8
        details_str = results.get('details', " ")
        # st.write("details_str:", details_str)
        detections = details_str.split('---')
        
        for detection in detections:
            if not detection.strip():
                continue  # Skip empty strings
            parts = detection.split('<br>')
            class_id, coordinates, confidence = None, None, None
            for part in parts:
                if 'Object type:' in part:
                    class_id_match = re.search(r"Object type:</b>\s*(\w+)", part)
                    if class_id_match:
                        class_id = class_id_match.group(1)
                elif 'Coordinates:' in part:
                    coordinates_match = re.search(r"Coordinates:</b>\s*\[(.*?)\]", part)
                    if coordinates_match:
                        coordinates = coordinates_match.group(1)
                elif 'Confidence:' in part:
                    confidence_match = re.search(r'Confidence:</b>\s*(\d+(\.\d+)?)', part)
                    if confidence_match:
                        confidence = float(confidence_match.group(1))

            if class_id and coordinates and confidence is not None:
                if class_id not in detailed_results:
                    detailed_results[class_id] = []
                detailed_results[class_id].append({
                    'model': model_name,
                    'coordinates': coordinates,
                    'confidence': confidence
                })
        
    # st.write("aggregated results:", aggregated_results, "\nDetailed results :", detailed_results)
    return aggregated_results, detailed_results

    # # Check if the model is YOLOv7
    # if model_name in config.DETECTION_MODEL_LIST_V7:
    #     count_dict = results.get(model_name, {}).get('count', {})
    #     if isinstance(count_dict, dict):
    #         for class_id, count_details in count_dict.items():
    #             if class_id not in aggregated_results:
    #                 aggregated_results[class_id] = {}
    #             aggregated_results[class_id][model_name] = count_details['count']
    #     elif isinstance(count_dict, int):
    #         # Handle the case where 'count' is an integer
    #         class_id = 'Unknown'
    #         if class_id not in aggregated_results:
    #             aggregated_results[class_id] = {}
    #         aggregated_results[class_id][model_name] = count_dict
    #     else:
    #         st.warning(f"Unexpected 'count' format for model {model_name}: {count_dict}")

    #     details_str = results.get(model_name, {}).get('details', {})
    #     if isinstance(details_str, str):
    #         detections = details_str.split('---')
    #         for detection in detections:
    #             parts = detection.split('<br>')
    #             class_id, coordinates, confidence = None, None, None
    #             for part in parts:
    #                 if 'Object type:' in part:
    #                     class_id_match = re.search(r"Object type:</b>\s*(\w+)", part)
    #                     if class_id_match:
    #                         class_id = class_id_match.group(1)
    #                     else:
    #                         print("Error parsing class ID:", part)
    #                         class_id = None
    #                 elif 'Coordinates:' in part:
    #                     coordinates_match = re.search(r"Coordinates:</b>\s*(.+)", part)
    #                     if coordinates_match:
    #                         coordinates = coordinates_match.group(1)
    #                     else:
    #                         print("Error parsing coordinates:", part)
    #                         coordinates = None
    #                 elif 'Confidence:' in part:
    #                     match = re.search(r'(\d+(\.\d+)?)%', part)
    #                     if match:
    #                         confidence = float(match.group(1))
    #                     else:
    #                         print("Error parsing confidence value:", part)
    #                         confidence = None

    #             if class_id and coordinates and confidence is not None:
    #                 if class_id not in detailed_results:
    #                     detailed_results[class_id] = []
    #                 detailed_results[class_id].append({
    #                     'model': model_name,
    #                     'coordinates': coordinates,
    #                     'confidence': confidence
    #                 })
    #     else:
    #         st.warning(f"Unexpected 'details' format for model {model_name}: {details_str}")

    # # Check if the model is YOLOv8
    # elif model_name in config.DETECTION_MODEL_LIST_V8:
    #     count = results.get("count", {})
    #     if isinstance(count, dict):
    #         for class_id, count_value in count.items():
    #             if class_id not in aggregated_results:
    #                 aggregated_results[class_id] = {}
    #             aggregated_results[class_id][model_name] = count_value
    #     elif isinstance(count, int):
    #         # Handle the case where 'count' is an integer
    #         class_id = 'Unknown'
    #         if class_id not in aggregated_results:
    #             aggregated_results[class_id] = {}
    #         aggregated_results[class_id][model_name] = count
    #     else:
    #         st.warning(f"Unexpected 'count' format for model {model_name}: {count}")

    #     details_str = results.get('details', " ")
    #     if isinstance(details_str, str):
    #         detections = details_str.split('---')
    #         for detection in detections:
    #             if not detection.strip():
    #                 continue
    #             parts = detection.split('<br>')
    #             class_id, coordinates, confidence = None, None, None
    #             for part in parts:
    #                 if 'Object type:' in part:
    #                     class_id_match = re.search(r"Object type:</b>\s*(\w+)", part)
    #                     if class_id_match:
    #                         class_id = class_id_match.group(1)
    #                 elif 'Coordinates:' in part:
    #                     coordinates_match = re.search(r"Coordinates:</b>\s*$$(.*?)$$", part)
    #                     if coordinates_match:
    #                         coordinates = coordinates_match.group(1)
    #                 elif 'Confidence:' in part:
    #                     confidence_match = re.search(r'Confidence:</b>\s*(\d+(\.\d+)?)', part)
    #                     if confidence_match:
    #                         confidence = float(confidence_match.group(1))

    #             if class_id and coordinates and confidence is not None:
    #                 if class_id not in detailed_results:
    #                     detailed_results[class_id] = []
    #                 detailed_results[class_id].append({
    #                     'model': model_name,
    #                     'coordinates': coordinates,
    #                     'confidence': confidence
    #                 })
    #     else:
    #         st.warning(f"Unexpected 'details' format for model {model_name}: {details_str}")

    # else:
    #     st.warning(f"Unexpected model name: {model_name}")

    # return aggregated_results, detailed_results


    


